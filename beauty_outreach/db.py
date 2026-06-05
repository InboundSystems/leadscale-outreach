import sqlite3
from datetime import date as date_type
from contextlib import contextmanager

import openpyxl

from .config import DB_PATH

SEED_CONTACTS = [
    (
        "Glamour Studio",
        "hello@glamourstudio.example.com",
        "Partnership Opportunity for Glamour Studio",
        "Hi Glamour Studio team,\n\nI came across your beautiful work and would love to explore a collaboration. We specialize in connecting top beauty brands with talented studios like yours.\n\nWould you be open to a quick call this week?\n\nBest,\nSam",
    ),
    (
        "Luxe Beauty Bar",
        "info@luxebeautybar.example.com",
        "Exclusive Offer for Luxe Beauty Bar",
        "Hello Luxe Beauty Bar,\n\nYour reputation for premium services caught our attention. We have an exclusive partnership program designed for high-end beauty businesses like yours.\n\nLet me know if you'd like to learn more!\n\nWarm regards,\nSam",
    ),
    (
        "The Brow Boutique",
        "contact@browboutique.example.com",
        "Quick Question for The Brow Boutique",
        "Hi there,\n\nI noticed The Brow Boutique has been growing rapidly — congratulations! I wanted to reach out about a growth opportunity that's been working well for similar boutiques.\n\nAre you free for a 15-minute chat?\n\nCheers,\nSam",
    ),
]


@contextmanager
def get_conn():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    try:
        yield conn
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def init_db():
    with get_conn() as conn:
        conn.executescript("""
            CREATE TABLE IF NOT EXISTS contacts (
                id INTEGER PRIMARY KEY,
                business_name TEXT,
                email TEXT UNIQUE,
                email_subject TEXT,
                unique_email_body TEXT,
                status TEXT DEFAULT 'queued',
                sequence_step INTEGER DEFAULT 0,
                last_sent_at TIMESTAMP,
                opened INTEGER DEFAULT 0,
                reply_received INTEGER DEFAULT 0,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );

            CREATE TABLE IF NOT EXISTS sent_log (
                id INTEGER PRIMARY KEY,
                contact_id INTEGER REFERENCES contacts(id),
                sequence_step INTEGER,
                sent_at TIMESTAMP,
                subject TEXT,
                opened_at TIMESTAMP,
                replied_at TIMESTAMP
            );

            CREATE TABLE IF NOT EXISTS warmup_log (
                id INTEGER PRIMARY KEY,
                sent_at TIMESTAMP,
                to_email TEXT,
                replied INTEGER DEFAULT 0
            );

            CREATE TABLE IF NOT EXISTS daily_stats (
                date TEXT PRIMARY KEY,
                emails_sent INTEGER DEFAULT 0,
                opens INTEGER DEFAULT 0,
                replies INTEGER DEFAULT 0,
                bounces INTEGER DEFAULT 0
            );
        """)

        for business_name, email, subject, body in SEED_CONTACTS:
            try:
                conn.execute(
                    """
                    INSERT INTO contacts (business_name, email, email_subject, unique_email_body)
                    VALUES (?, ?, ?, ?)
                    """,
                    (business_name, email, subject, body),
                )
            except sqlite3.IntegrityError:
                pass


def import_contacts_from_excel(filepath: str) -> dict:
    wb = openpyxl.load_workbook(filepath, data_only=True)

    sheet_configs = [
        "2 - Email + Phone",
        "3 - Email Only",
    ]

    # Column indices (1-based as openpyxl uses 1-based by default via cell(row, col))
    COL_BUSINESS_NAME = 2
    COL_EMAIL = 4
    COL_SUBJECT = 12
    COL_BODY = 13

    imported = 0
    updated = 0
    skipped_already_sent = 0
    skipped_empty = 0

    with get_conn() as conn:
        for sheet_name in sheet_configs:
            if sheet_name not in wb.sheetnames:
                continue

            ws = wb[sheet_name]

            for row in ws.iter_rows(min_row=3, values_only=True):
                business_name = row[COL_BUSINESS_NAME - 1]
                email = row[COL_EMAIL - 1]
                subject = row[COL_SUBJECT - 1]
                body = row[COL_BODY - 1]

                if not email or str(email).strip() == "":
                    skipped_empty += 1
                    continue

                if not body or str(body).strip() == "":
                    skipped_empty += 1
                    continue

                email = str(email).strip()
                business_name = str(business_name).strip() if business_name else ""
                subject = str(subject).strip() if subject else ""
                body = str(body).strip()

                try:
                    conn.execute(
                        """
                        INSERT INTO contacts (business_name, email, email_subject, unique_email_body)
                        VALUES (?, ?, ?, ?)
                        """,
                        (business_name, email, subject, body),
                    )
                    imported += 1
                except sqlite3.IntegrityError:
                    # Contact exists — update subject/body only if not yet sent
                    existing = conn.execute(
                        "SELECT status FROM contacts WHERE email = ?", (email,)
                    ).fetchone()
                    if existing and existing["status"] == "queued":
                        conn.execute(
                            """
                            UPDATE contacts
                            SET business_name = ?, email_subject = ?, unique_email_body = ?
                            WHERE email = ?
                            """,
                            (business_name, subject, body, email),
                        )
                        updated += 1
                    else:
                        skipped_already_sent += 1

    return {
        "imported": imported,
        "updated": updated,
        "skipped_already_sent": skipped_already_sent,
        "skipped_empty": skipped_empty,
    }


def init_batch3_schedule():
    with get_conn() as conn:
        conn.executescript("""
            CREATE TABLE IF NOT EXISTS batch3_schedule (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                master_day INTEGER,
                scheduled_date TEXT,
                business_name TEXT,
                email TEXT,
                suburb TEXT,
                category TEXT,
                email_step TEXT,
                lead_score INTEGER,
                business_type TEXT,
                email_subject TEXT,
                email_body TEXT,
                status TEXT DEFAULT 'scheduled',
                sent_at TEXT,
                reply_received INTEGER DEFAULT 0
            );
            CREATE INDEX IF NOT EXISTS idx_batch3_date  ON batch3_schedule(scheduled_date);
            CREATE INDEX IF NOT EXISTS idx_batch3_email ON batch3_schedule(email);
        """)


def import_batch3_from_excel(filepath: str) -> dict:
    from datetime import datetime as dt
    init_batch3_schedule()
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb['📧 Email Schedule']

    imported = skipped = 0

    with get_conn() as conn:
        conn.execute("DELETE FROM batch3_schedule WHERE status = 'scheduled'")

        for row in range(2, ws.max_row + 1):
            master_day = ws.cell(row, 1).value
            if not isinstance(master_day, int):
                continue

            date_raw     = ws.cell(row, 2).value
            business_name = ws.cell(row, 4).value
            email        = ws.cell(row, 5).value
            suburb       = ws.cell(row, 6).value
            category     = ws.cell(row, 7).value
            email_step   = ws.cell(row, 8).value
            lead_score   = ws.cell(row, 9).value
            business_type = ws.cell(row, 10).value
            email_subject = ws.cell(row, 11).value
            email_body   = ws.cell(row, 12).value

            if not email:
                skipped += 1
                continue

            email = str(email).strip()

            if isinstance(date_raw, str):
                try:
                    scheduled_date = dt.strptime(date_raw, '%d/%m/%Y').strftime('%Y-%m-%d')
                except ValueError:
                    skipped += 1
                    continue
            elif hasattr(date_raw, 'strftime'):
                scheduled_date = date_raw.strftime('%Y-%m-%d')
            else:
                skipped += 1
                continue

            conn.execute("""
                INSERT INTO batch3_schedule
                (master_day, scheduled_date, business_name, email, suburb, category,
                 email_step, lead_score, business_type, email_subject, email_body)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                master_day, scheduled_date,
                str(business_name).strip() if business_name else '',
                email,
                str(suburb).strip() if suburb else '',
                str(category).strip() if category else '',
                str(email_step).strip() if email_step else '',
                lead_score,
                str(business_type).strip() if business_type else '',
                str(email_subject).strip() if email_subject else None,
                str(email_body).strip() if email_body else None,
            ))
            imported += 1

    return {'imported': imported, 'skipped': skipped}


def get_batch3_emails_for_date(date: str) -> list:
    init_batch3_schedule()
    with get_conn() as conn:
        return conn.execute("""
            SELECT * FROM batch3_schedule
            WHERE scheduled_date = ? AND status = 'scheduled'
            ORDER BY id ASC
        """, (date,)).fetchall()


def was_batch3_initial_sent(email: str) -> bool:
    with get_conn() as conn:
        row = conn.execute("""
            SELECT id FROM batch3_schedule
            WHERE email = ? AND email_step = 'Initial Email' AND status = 'sent'
        """, (email,)).fetchone()
    return row is not None


def mark_batch3_sent(record_id: int, sent_at: str):
    with get_conn() as conn:
        conn.execute(
            "UPDATE batch3_schedule SET status = 'sent', sent_at = ? WHERE id = ?",
            (sent_at, record_id)
        )


def mark_batch3_skipped(record_id: int):
    with get_conn() as conn:
        conn.execute(
            "UPDATE batch3_schedule SET status = 'skipped' WHERE id = ?",
            (record_id,)
        )


def mark_batch3_bounced(email: str):
    with get_conn() as conn:
        conn.execute(
            "UPDATE batch3_schedule SET status = 'bounced' WHERE email = ? AND status = 'scheduled'",
            (email,)
        )


def mark_batch3_reply(email: str):
    with get_conn() as conn:
        conn.execute(
            "UPDATE batch3_schedule SET reply_received = 1 WHERE email = ?",
            (email,)
        )


def get_daily_send_count(date: str) -> int:
    with get_conn() as conn:
        row = conn.execute(
            "SELECT emails_sent FROM daily_stats WHERE date = ?", (date,)
        ).fetchone()
    return row["emails_sent"] if row else 0


def update_daily_stats(date: str, field: str, increment_by: int = 1):
    allowed_fields = {"emails_sent", "opens", "replies", "bounces"}
    if field not in allowed_fields:
        raise ValueError(f"Invalid field: {field}. Must be one of {allowed_fields}")

    with get_conn() as conn:
        conn.execute(
            f"""
            INSERT INTO daily_stats (date, {field})
            VALUES (?, ?)
            ON CONFLICT(date) DO UPDATE SET {field} = {field} + excluded.{field}
            """,
            (date, increment_by),
        )
